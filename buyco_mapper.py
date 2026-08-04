#!/usr/bin/env python3
"""Map harvested TariffDoc records (from dnd_tariff_scraper.py) into BuyCo's
D&D contract-rule template ("Template (6).xlsx" - Sheet1 schema, with
Sheet2/Sheet3 lookups preserved).

WHAT THIS DOES:
  For every discovered tariff document, infer (Type, Penalty type, free
  time, tiered per-diem rates, currency, port) from the document's title
  and extracted PDF text/tables, and emit one row per (direction x penalty
  type x port) combination the document plausibly covers.

WHAT THIS DELIBERATELY DOES NOT DO:
  Silently guess a single "best" answer when the source is ambiguous. Where
  a field can't be determined with confidence, every plausible value is
  emitted as a separate row (e.g. both Origin and Destination) and flagged
  in the "QA Notes" column (Sheet1 col Z) plus collected on a dedicated
  "Needs Review" sheet, rather than picking one silently.

STATUS: none of the extraction regexes below have been run against a real
carrier tariff PDF - this session had no network access to fetch one. They
were designed from the template's own examples and general knowledge of
how these tariffs are usually worded, then unit-tested against fixtures
that mimic that wording. Treat every row this produces as a draft: check a
handful against the source PDF before trusting the numbers.

Known caveats, mirrored from the BuyCo-template Q&A:
  - Contract number -> CONTRACT_NUMBER_FLAG ("SCRAPED - STANDARD TARIFF")
  - Shipper -> blank (applies to everyone)
  - Contract start/end date -> the tariff's own effective/expiration date,
    not a BuyCo contract window (end date is legitimately blank when the
    source doesn't publish one - most don't)
  - POL/POD/POR/PODEL -> single named port when the source specifies one;
    otherwise fanned out across ports_reference.PORTS_BY_COUNTRY (itself
    unverified - see that file's docstring)
  - D&D vs split Demurrage/Detention -> follows whatever the source text
    actually contains
  - Penalty 2/3 tiers -> left blank when the source has fewer tiers
  - Validity milestone -> defaulted to "ETD@EPU" (matches every example row
    in the template) since nothing in a carrier's published tariff
    indicates which BuyCo milestone should anchor validity - flagged
    accordingly
  - Date calculation -> defaulted to "Day Calendar" (matches every example
    row) for the same reason
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field as dc_field
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

try:
    from dateutil import parser as dateutil_parser
except ImportError:
    dateutil_parser = None

from ports_reference import get_ports_for_country

try:
    from dnd_tariff_scraper import TariffDoc, CONTRACT_NUMBER_FLAG
except ImportError:  # allow standalone import/testing without the scraper module
    CONTRACT_NUMBER_FLAG = "SCRAPED - STANDARD TARIFF"
    TariffDoc = object  # type: ignore

NEEDS_REVIEW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

DEFAULT_VALIDITY_MILESTONE = "ETD@EPU"
DEFAULT_DATE_CALC = "Day Calendar"


# ---------------------------------------------------------------------------
# Row model
# ---------------------------------------------------------------------------

@dataclass
class MappedRow:
    type_: str                 # "Origin" | "Destination"
    container_type: str = ""
    hazardous: str = "F"
    carrier: str = ""
    shipper: str = ""
    por: str = ""
    pol: str = ""
    pod: str = ""
    podel: str = ""
    contract_number: str = CONTRACT_NUMBER_FLAG
    contract_start: str = ""
    contract_end: str = ""
    validity_milestone: str = DEFAULT_VALIDITY_MILESTONE
    penalty_type: str = ""
    currency: str = ""
    date_calc: str = DEFAULT_DATE_CALC
    free_time: Optional[int] = None
    penalty1: Optional[float] = None
    period1: Optional[int] = None
    penalty2: Optional[float] = None
    period2: Optional[int] = None
    penalty3: Optional[float] = None
    qa_notes: list[str] = dc_field(default_factory=list)
    source_title: str = ""
    source_pdf: str = ""

    def event1_event2(self) -> tuple[str, str]:
        return compute_events(self.type_, self.penalty_type)


# ---------------------------------------------------------------------------
# Extraction heuristics (operate on doc.title / doc.raw_text)
# ---------------------------------------------------------------------------

FREE_TIME_RE = re.compile(
    r"free\s*time[^\d]{0,20}(\d{1,3})\s*(?:calendar\s*|working\s*)?days?"
    r"|(\d{1,3})\s*(?:calendar\s*|working\s*)?days?\s*(?:of\s*)?free\s*time",
    re.IGNORECASE,
)

CURRENCY_RE = re.compile(r"\b(USD|EUR|GBP|CNY|SGD|AED|JPY|INR|CAD|AUD|MXN|BRL)\b")

# "day(s) 1-5 ... USD 50" / "1-5 days ... $50" / "1st to 5th day ... $50"
# style tier lines - the "days" word can land before or after the range.
TIER_RE = re.compile(
    r"days?\s*(\d{1,3})\s*(?:st|nd|rd|th)?\s*(?:-|to|–)\s*(\d{1,3})\s*(?:st|nd|rd|th)?"
    r"[^\d\n]{0,30}?(?:USD|US\$|\$|EUR|€|GBP|£)?\s*([\d,]+(?:\.\d+)?)"
    r"|(\d{1,3})\s*(?:st|nd|rd|th)?\s*(?:-|to|–)\s*(\d{1,3})\s*(?:st|nd|rd|th)?\s*days?"
    r"[^\d\n]{0,30}?(?:USD|US\$|\$|EUR|€|GBP|£)?\s*([\d,]+(?:\.\d+)?)",
    re.IGNORECASE,
)

# Flat/no-tier rate: "USD 50 per day" with no day range attached.
FLAT_RATE_RE = re.compile(
    r"(?:USD|US\$|\$|EUR|€|GBP|£)\s*([\d,]+(?:\.\d+)?)\s*(?:per\s*day|/\s*day|daily)",
    re.IGNORECASE,
)

# Open-ended final tier: "day 11 onwards/thereafter/and beyond ... USD 150".
# No end day, so period is left None (an indefinite last tier, matching the
# template's own examples where the final Penalty column has no Period pair).
OPEN_ENDED_TIER_RE = re.compile(
    r"days?\s*(\d{1,3})\s*(?:onwards|and\s+beyond|thereafter|\+)"
    r"[^\d\n]{0,30}?(?:USD|US\$|\$|EUR|€|GBP|£)?\s*([\d,]+(?:\.\d+)?)",
    re.IGNORECASE,
)


def extract_free_time(text: str) -> Optional[int]:
    m = FREE_TIME_RE.search(text or "")
    if not m:
        return None
    val = m.group(1) or m.group(2)
    return int(val)


def extract_currency(text: str) -> str:
    m = CURRENCY_RE.search(text or "")
    return m.group(1).upper() if m else ""


def extract_tiers(text: str) -> list[tuple[float, Optional[int]]]:
    """Returns up to 3 (rate, period_days) tuples, in document order."""
    tiers: list[tuple[float, Optional[int]]] = []
    for m in TIER_RE.finditer(text or ""):
        g = m.groups()
        start, end, amount = (g[0], g[1], g[2]) if g[0] is not None else (g[3], g[4], g[5])
        rate = float(amount.replace(",", ""))
        period = max(int(end) - int(start) + 1, 1)
        tiers.append((rate, period))
        if len(tiers) == 3:
            break
    if tiers and len(tiers) < 3:
        m = OPEN_ENDED_TIER_RE.search(text or "")
        if m:
            tiers.append((float(m.group(2).replace(",", "")), None))
    if tiers:
        return tiers[:3]
    m = FLAT_RATE_RE.search(text or "")
    if m:
        return [(float(m.group(1).replace(",", "")), None)]
    return []


PENALTY_KEYWORDS = ["D&D", "Demurrage", "Detention", "Storage"]


def extract_penalty_types(title: str, raw_text: str) -> list[str]:
    """Which penalty type(s) this document plausibly covers.

    If the title names exactly one, trust that. If it names more than one
    (or none, and the body text mentions more than one), surface all of
    them as separate rows rather than merging - per the "follow the source"
    call, splitting is only correct once we can see whether the source
    actually gives them separate free-time/rate figures, which requires a
    real sample; until then, over-surfacing is safer than silently merging.
    """
    haystack = f"{title}\n{raw_text or ''}"
    found = [kw for kw in PENALTY_KEYWORDS if kw.lower() in haystack.lower()]
    if not found:
        return ["D&D"]  # most generic/common default; always flagged by caller
    # "D&D" alongside Demurrage/Detention is redundant phrasing, not a 3rd type
    if "D&D" in found and ("Demurrage" in found or "Detention" in found):
        found = [f for f in found if f != "D&D"]
    return found


def extract_directions(title: str) -> list[str]:
    lowered = title.lower()
    has_import = "import" in lowered
    has_export = "export" in lowered
    if has_import and not has_export:
        return ["Destination"]
    if has_export and not has_import:
        return ["Origin"]
    if has_import and has_export:
        return ["Origin", "Destination"]
    return ["Origin", "Destination"]  # unknown - surface both, flagged by caller


# ---------------------------------------------------------------------------
# Event1/Event2 - direct transcription of the template's R2/S2 formulas
# ---------------------------------------------------------------------------

def compute_events(direction: str, penalty_type: str) -> tuple[str, str]:
    if direction == "Destination" and penalty_type in ("D&D", "Demurrage", "Storage"):
        e1 = "DISCHARGE"
    elif direction == "Destination" and penalty_type == "Detention":
        e1 = "GATE_OUT"
    elif direction == "Origin" and penalty_type in ("D&D", "Detention"):
        e1 = "EMPTY_PICKUP"
    elif direction == "Origin" and penalty_type in ("Demurrage", "Storage"):
        e1 = "GATE_IN"
    else:
        e1 = ""

    if direction == "Destination" and penalty_type in ("Demurrage", "Storage"):
        e2 = "GATE_OUT"
    elif direction == "Destination" and penalty_type in ("Detention", "D&D"):
        e2 = "EMPTY_RETURN"
    elif direction == "Origin" and penalty_type == "Detention":
        e2 = "GATE_IN"
    elif direction == "Origin" and penalty_type in ("Demurrage", "Storage", "D&D"):
        e2 = "LOADING"
    else:
        e2 = ""
    return e1, e2


# ---------------------------------------------------------------------------
# Port resolution
# ---------------------------------------------------------------------------

def resolve_ports(country: str, raw_text: str) -> tuple[list[str], list[str]]:
    """Returns (unlocodes, notes). notes explains how they were chosen."""
    ref_ports = get_ports_for_country(country)
    if not ref_ports:
        return [], [f"No port reference entry for country '{country}' - "
                     f"POL/POD left blank, needs manual port mapping."]

    if raw_text:
        named = [code for code, name in ref_ports
                 if name.split("/")[0].strip().lower() in raw_text.lower()]
        if named:
            return named, [f"Port(s) matched by name in source text: {named}."]

    codes = [code for code, _ in ref_ports]
    return codes, [f"Source doesn't name a specific port - fanned out across "
                    f"{len(codes)} reference port(s) for {country} "
                    f"(best-effort list, not exhaustive)."]


# ---------------------------------------------------------------------------
# CMA CGM - structured parser (calibrated against real fetched PDFs: France
# and US, 2026-08-04). CMA's tariff PDFs follow a stable nested template:
#
#   D&D TARIFFS <COUNTRY>
#   IMPORT|EXPORT - <PLACE DESCRIPTION> [(UNLOCODE)]
#   TARIFF IN <CURRENCY WORD, e.g. EURO/USD>
#   ...
#   EFFECTIVE DATE*: <DD-MON-YYYY>
#   EXPIRATION DATE: <DD-MON-YYYY | UNTIL FURTHER NOTICE>
#   ...
#   <STANDARD|NOR CONTAINER|REEFER CONTAINER|SPECIAL CONTAINER|HAZARDOUS|TK> - SPLITTED
#   DEMURRAGE [DETENTION] | DETENTION | STORAGE
#   SLAB/TIER (in days) 20' 40' 45' [SLAB/TIER (in days) 20' 40' 45']
#   <N> FREE DAY(S) [<N> FREE DAY(S)]
#   From <ord> To <ord> <r20> <r40> <r45> [From <ord> To <ord> <r20> <r40> <r45>]
#   From <ord> Onwards <r20> <r40> <r45> [...]
#   ... (repeats for each container-type block, then repeats for each place)
#
# One PDF can contain many place-blocks (one per port/region) and each
# place-block can contain several container-type blocks, each of which can
# cover Demurrage and Detention side by side on the same lines. This walks
# the text top-to-bottom as a small state machine instead of one flat
# regex, since the structure genuinely nests three levels deep and a flat
# regex can't tell which numbers belong to which (port, container
# type/size, penalty type) combination.
#
# NOT yet verified against every CMA country - only France and US. Other
# countries may phrase tiers slightly differently (e.g. "Until Day X"
# instead of "To Xth"); anything this can't parse is flagged in QA Notes
# rather than silently dropped, so gaps stay visible.
# ---------------------------------------------------------------------------

CMA_PLACE_HEADER_RE = re.compile(r"^\s*(IMPORT|EXPORT)\s*[-–]\s*(.+?)\s*$", re.IGNORECASE)
CMA_CONTAINER_HEADER_RE = re.compile(
    r"^\s*(STANDARD|NOR CONTAINER|REEFER CONTAINER|SPECIAL CONTAINER|HAZARDOUS|TK)\s*-\s*SPLITTED\s*$",
    re.IGNORECASE,
)
CMA_PENALTY_HEADER_RE = re.compile(
    r"^\s*(DEMURRAGE|DETENTION|STORAGE)(?:\s+(DEMURRAGE|DETENTION|STORAGE))?\s*$", re.IGNORECASE
)
CMA_CURRENCY_RE = re.compile(r"TARIFF IN\s+([A-Z]+)", re.IGNORECASE)
CMA_EFFECTIVE_RE = re.compile(r"EFFECTIVE DATE\*?:\s*([\d]{1,2}[-\s][A-Za-z]{3}[-\s][\d]{2,4})", re.IGNORECASE)
CMA_EXPIRATION_RE = re.compile(r"EXPIRATION DATE:\s*(.+)", re.IGNORECASE)
CMA_NAMED_PORT_RE = re.compile(r"\(([A-Z]{5})\)")
CMA_FREE_DAYS_RE = re.compile(r"(\d{1,3})\s*(?:CALENDAR\s+)?FREE\s+DAYS?", re.IGNORECASE)
CMA_TIER_FRAGMENT_RE = re.compile(
    r"From\s+(\d{1,3})(?:st|nd|rd|th)?\s+(?:To\s+(\d{1,3})(?:st|nd|rd|th)?|Onwards)\s+"
    r"([\d.]+)\s+([\d.]+)\s+([\d.]+)",
    re.IGNORECASE,
)

CMA_CURRENCY_ALIASES = {"EURO": "EUR", "USED": "USD"}

# Matches BuyCo's actual Sheet3 taxonomy (DV/HC/RFR, 20DV/40DV/40HC/20RFR/40RFR
# - no 45' entries). Per BuyCo's call:
#   - 45' is dropped entirely (only indices 0/1 -> 20'/40' are ever read;
#     whatever the source's 45' column says is discarded, even when it
#     differs from 40', e.g. some NY/NJ lines)
#   - NOR CONTAINER (a reefer run as a standard box, per the source's own
#     description) maps to DV codes, not RFR
#   - SPECIAL CONTAINER (open tops/flat racks) and TK (tanks) both map to
#     TK codes, sized (20TK/40TK isn't in Sheet3 yet - add it there, or
#     tell me to rename these strings, if it should be unsized "TK" instead)
CMA_CONTAINER_SIZE_CODES = {
    "STANDARD": ["20DV", "40DV"],
    "NOR CONTAINER": ["20DV", "40DV"],
    "REEFER CONTAINER": ["20RFR", "40RFR"],
    "SPECIAL CONTAINER": ["20TK", "40TK"],
    "HAZARDOUS": ["20DV", "40DV"],
    "TK": ["20TK", "40TK"],
}


def parse_cma_cgm(doc) -> list[MappedRow]:
    raw_text = getattr(doc, "raw_text", "") or ""
    carrier = getattr(doc, "carrier", "") or ""
    country = getattr(doc, "country", "") or ""
    pdf_url = getattr(doc, "pdf_url", "") or ""
    title = getattr(doc, "title", "") or ""
    lines = raw_text.splitlines()

    # Pre-scan: every UNLOCODE named anywhere in this doc gets its own
    # place-block already, so a later "ALL PORTS EXCEPT..." fanout
    # shouldn't duplicate them with a second, more generic row.
    all_named_ports = set(CMA_NAMED_PORT_RE.findall(raw_text))

    rows: list[MappedRow] = []

    direction = None
    place_desc = ""
    named_port = ""
    currency = ""
    effective = getattr(doc, "effective_date_guess", "") or ""
    expiration = getattr(doc, "validity_end_guess", "") or ""

    container_header = ""
    penalty_types: list[str] = []
    free_days: dict[str, int] = {}
    tiers: dict[str, list[tuple[float, float, float, Optional[int]]]] = {}

    def flush_container_block():
        if not penalty_types or direction is None:
            return
        codes = CMA_CONTAINER_SIZE_CODES.get(container_header, ["20DV", "40DV"])
        hazardous = "T" if container_header == "HAZARDOUS" else "F"

        if named_port:
            target_ports = [named_port]
            port_notes = []
        else:
            ref_ports = [c for c, _ in get_ports_for_country(country) if c not in all_named_ports]
            if ref_ports:
                target_ports = ref_ports
                port_notes = [f"'{place_desc}' doesn't name a specific port - fanned out across "
                              f"{len(ref_ports)} reference port(s) for {country}, excluding any "
                              f"port already named elsewhere in this same PDF."]
            else:
                target_ports = [""]
                port_notes = [f"'{place_desc}' doesn't name a specific port and no usable port "
                              f"reference for {country} - POL/POD left blank, needs manual mapping."]

        for penalty_type in penalty_types:
            ft = free_days.get(penalty_type)
            tier_list = tiers.get(penalty_type, [])
            base_notes = list(port_notes)
            if ft is None:
                base_notes.append(f"No free-day count parsed for {penalty_type} in this block.")
            if not tier_list:
                base_notes.append(f"No rate tiers parsed for {penalty_type} in this block.")
            for target_port in target_ports:
                for size_idx, code in enumerate(codes):
                    row = MappedRow(
                        type_=direction,
                        container_type=code,
                        hazardous=hazardous,
                        carrier=carrier,
                        penalty_type=penalty_type,
                        currency=currency,
                        free_time=ft,
                        contract_start=effective,
                        contract_end="" if expiration.upper() == "UNTIL FURTHER NOTICE" else expiration,
                        qa_notes=list(base_notes),
                        source_title=title,
                        source_pdf=pdf_url,
                    )
                    if direction == "Destination":
                        row.pod = target_port
                    else:
                        row.pol = target_port
                    for tier_i, tier in enumerate(tier_list[:3]):
                        r20, r40, r45, period = tier
                        rate = {0: r20, 1: r40, 2: r45}[size_idx]
                        if tier_i == 0:
                            row.penalty1, row.period1 = rate, period
                        elif tier_i == 1:
                            row.penalty2, row.period2 = rate, period
                        elif tier_i == 2:
                            row.penalty3 = rate
                    rows.append(row)

    for line in lines:
        line = line.strip()

        m = CMA_PLACE_HEADER_RE.match(line)
        if m:
            flush_container_block()
            container_header, penalty_types, free_days, tiers = "", [], {}, {}
            direction = "Destination" if m.group(1).upper() == "IMPORT" else "Origin"
            place_desc = m.group(2)
            port_m = CMA_NAMED_PORT_RE.search(place_desc)
            named_port = port_m.group(1) if port_m else ""
            continue

        cm = CMA_CURRENCY_RE.search(line)
        if cm:
            token = cm.group(1).upper()
            currency = CMA_CURRENCY_ALIASES.get(token, token)

        em = CMA_EFFECTIVE_RE.search(line)
        if em:
            effective = em.group(1)

        xm = CMA_EXPIRATION_RE.search(line)
        if xm:
            expiration = xm.group(1).strip()

        ch = CMA_CONTAINER_HEADER_RE.match(line)
        if ch:
            flush_container_block()
            container_header = ch.group(1).upper()
            penalty_types, free_days, tiers = [], {}, {}
            continue

        ph = CMA_PENALTY_HEADER_RE.match(line)
        if ph and container_header:
            penalty_types = [g.title() for g in ph.groups() if g]
            continue

        if penalty_types:
            free_matches = list(CMA_FREE_DAYS_RE.finditer(line))
            for idx, fm in enumerate(free_matches[:len(penalty_types)]):
                free_days[penalty_types[idx]] = int(fm.group(1))

            tier_matches = list(CMA_TIER_FRAGMENT_RE.finditer(line))
            for idx, tm in enumerate(tier_matches[:len(penalty_types)]):
                start, end = tm.group(1), tm.group(2)
                period = (int(end) - int(start) + 1) if end else None
                r20, r40, r45 = float(tm.group(3)), float(tm.group(4)), float(tm.group(5))
                tiers.setdefault(penalty_types[idx], []).append((r20, r40, r45, period))

    flush_container_block()

    if not rows:
        rows.append(MappedRow(
            type_="Origin", carrier=carrier, penalty_type="D&D",
            source_title=title, source_pdf=pdf_url,
            qa_notes=["parse_cma_cgm found no place/container-type blocks at all - this "
                      "document's wording doesn't match the France/US pattern this parser "
                      "was built against. Needs a look at the actual PDF."],
        ))
    return rows


# ---------------------------------------------------------------------------
# Main per-document mapper
# ---------------------------------------------------------------------------

def map_doc(doc) -> list[MappedRow]:
    if getattr(doc, "carrier", "") == "CMA CGM":
        return parse_cma_cgm(doc)
    title = getattr(doc, "title", "") or ""
    raw_text = getattr(doc, "raw_text", "") or ""
    country = getattr(doc, "country", "") or ""
    carrier = getattr(doc, "carrier", "") or ""

    directions = extract_directions(title)
    penalty_types = extract_penalty_types(title, raw_text)
    ports, port_notes = resolve_ports(country, raw_text)
    free_time = extract_free_time(raw_text or title)
    tiers = extract_tiers(raw_text or title)
    currency = extract_currency(raw_text or title)

    notes: list[str] = list(port_notes)
    if len(directions) > 1:
        notes.append("Direction (Origin vs Destination) not stated in the title - "
                      "both emitted, delete the one that doesn't apply.")
    if free_time is None:
        notes.append("Could not find a free-time day count in the extracted text.")
    if not tiers:
        notes.append("Could not find a per-diem rate/tier in the extracted text.")
    if not currency:
        notes.append("Could not detect a currency code in the extracted text.")
    if not ports:
        ports = [""]  # still emit one row with blank ports rather than dropping the doc

    rows: list[MappedRow] = []
    for direction in directions:
        for penalty_type in penalty_types:
            for port in ports:
                row = MappedRow(
                    type_=direction,
                    carrier=carrier,
                    penalty_type=penalty_type,
                    currency=currency,
                    free_time=free_time,
                    contract_start=getattr(doc, "effective_date_guess", "") or "",
                    contract_end=getattr(doc, "validity_end_guess", "") or "",
                    qa_notes=list(notes),
                    source_title=title,
                    source_pdf=getattr(doc, "pdf_url", "") or "",
                )
                if direction == "Destination":
                    row.pod = port
                else:
                    row.pol = port
                if tiers:
                    if len(tiers) >= 1:
                        row.penalty1, row.period1 = tiers[0]
                    if len(tiers) >= 2:
                        row.penalty2, row.period2 = tiers[1]
                    if len(tiers) >= 3:
                        row.penalty3, _ = tiers[2]
                rows.append(row)
    return rows


# ---------------------------------------------------------------------------
# Template writer
# ---------------------------------------------------------------------------

FORMULA_CELL_REF_RE = re.compile(r"\b([A-Z]{1,2})2\b")


def retarget_formula(formula: str, new_row: int) -> str:
    """Rewrites a row-2 formula (e.g. '=...D2...') to reference new_row instead."""
    return FORMULA_CELL_REF_RE.sub(lambda m: f"{m.group(1)}{new_row}", formula)


def write_template(template_path: Path, docs: list, out_path: Path) -> dict:
    """Clones the BuyCo template, appends mapped rows to Sheet1, and adds a
    'Needs Review' sheet. Returns a small summary dict."""
    wb = load_workbook(template_path)
    ws = wb["Sheet1"]

    # Grab the live formula patterns from row 2 so new rows stay consistent
    # with however BuyCo has this template wired, rather than hardcoding them.
    e_formula = ws["E2"].value if ws["E2"].value and str(ws["E2"].value).startswith("=") else None
    r_formula = ws["R2"].value if ws["R2"].value and str(ws["R2"].value).startswith("=") else None
    s_formula = ws["S2"].value if ws["S2"].value and str(ws["S2"].value).startswith("=") else None

    next_row = ws.max_row + 1
    review_rows = []
    total = 0

    for doc in docs:
        for mrow in map_doc(doc):
            r = next_row
            ws.cell(r, 1, mrow.type_)
            ws.cell(r, 2, mrow.container_type)
            ws.cell(r, 3, mrow.hazardous)
            ws.cell(r, 4, mrow.carrier)
            if e_formula:
                ws.cell(r, 5, retarget_formula(e_formula, r))
            ws.cell(r, 6, mrow.shipper)
            ws.cell(r, 7, mrow.por)
            ws.cell(r, 8, mrow.pol)
            ws.cell(r, 9, mrow.pod)
            ws.cell(r, 10, mrow.podel)
            ws.cell(r, 11, mrow.contract_number)
            ws.cell(r, 12, mrow.contract_start)
            ws.cell(r, 13, mrow.contract_end)
            ws.cell(r, 14, mrow.validity_milestone)
            ws.cell(r, 15, mrow.penalty_type)
            ws.cell(r, 16, mrow.currency)
            ws.cell(r, 17, mrow.date_calc)
            if r_formula:
                ws.cell(r, 18, retarget_formula(r_formula, r))
            if s_formula:
                ws.cell(r, 19, retarget_formula(s_formula, r))
            ws.cell(r, 20, mrow.free_time)
            ws.cell(r, 21, mrow.penalty1)
            ws.cell(r, 22, mrow.period1)
            ws.cell(r, 23, mrow.penalty2)
            ws.cell(r, 24, mrow.period2)
            ws.cell(r, 25, mrow.penalty3)

            qa_text = " | ".join(mrow.qa_notes)
            qa_cell = ws.cell(r, 26, qa_text)
            if mrow.qa_notes:
                for col in range(1, 27):
                    ws.cell(r, col).fill = NEEDS_REVIEW_FILL
                review_rows.append((r, mrow, qa_text))

            next_row += 1
            total += 1

    if ws.cell(1, 26).value != "QA Notes":
        ws.cell(1, 26, "QA Notes")

    review_ws = wb.create_sheet("Needs Review")
    review_ws.append(["Row in Sheet1", "Carrier", "Country/Type", "Penalty Type",
                       "Source Title", "Source PDF", "QA Notes"])
    for r, mrow, qa_text in review_rows:
        review_ws.append([r, mrow.carrier, mrow.type_, mrow.penalty_type,
                           mrow.source_title, mrow.source_pdf, qa_text])

    wb.save(out_path)
    return {
        "rows_written": total,
        "rows_flagged": len(review_rows),
        "output": str(out_path),
    }


# ---------------------------------------------------------------------------
# Monthly upsert - match key = Carrier + Direction + Port + Penalty type.
# On match: only Contract end date is refreshed (and only if the new run's
# guess parses to a later date than what's there). Rows flagged "Manually
# Reviewed" are never touched, matching or not.
# ---------------------------------------------------------------------------

MANUALLY_REVIEWED_COL = 27  # column AA


def parse_date_best_effort(value) -> Optional["__import__('datetime').date"]:
    if not value:
        return None
    if hasattr(value, "date"):  # already a datetime/date from openpyxl
        return value.date() if hasattr(value, "year") and hasattr(value, "hour") else value
    if dateutil_parser is None:
        return None
    try:
        return dateutil_parser.parse(str(value), fuzzy=True).date()
    except (ValueError, OverflowError):
        return None


def row_key(mrow: "MappedRow") -> tuple:
    # Container type is part of the identity now that CMA rows are split per
    # size/equipment type (20DV vs 40DV vs 20RFR, etc.) - without it, two
    # rows for the same port/penalty that only differ by size would look
    # identical to the matcher and collide.
    port = mrow.pod if mrow.type_ == "Destination" else mrow.pol
    return (mrow.carrier, mrow.type_, port, mrow.penalty_type, mrow.container_type)


def upsert_template(template_path: Path, docs: list, out_path: Path) -> dict:
    """Same row-building as write_template, but merges into an EXISTING
    workbook (e.g. downloaded fresh from Drive) instead of always appending.

    - New (carrier, direction, port, penalty type) combo -> append, same as
      write_template.
    - Existing combo, not marked Manually Reviewed -> update Contract end
      date only if the new run found a later one; nothing else changes.
    - Existing combo, marked Manually Reviewed (col AA non-empty) -> skipped
      entirely, no exceptions.
    """
    wb = load_workbook(template_path)
    ws = wb["Sheet1"]

    if ws.cell(1, 26).value != "QA Notes":
        ws.cell(1, 26, "QA Notes")
    if ws.cell(1, MANUALLY_REVIEWED_COL).value != "Manually Reviewed":
        ws.cell(1, MANUALLY_REVIEWED_COL, "Manually Reviewed")

    e_formula = ws["E2"].value if ws["E2"].value and str(ws["E2"].value).startswith("=") else None
    r_formula = ws["R2"].value if ws["R2"].value and str(ws["R2"].value).startswith("=") else None
    s_formula = ws["S2"].value if ws["S2"].value and str(ws["S2"].value).startswith("=") else None

    # Build lookup of existing rows by key, INCLUDING Manually Reviewed ones -
    # they still need to be matchable so a future run recognizes "this key
    # already has a row" and skips outright, rather than not finding it and
    # appending a silent duplicate next to the one a human already fixed.
    existing_index: dict[tuple, int] = {}
    for r in range(2, ws.max_row + 1):
        type_ = ws.cell(r, 1).value
        container_type = ws.cell(r, 2).value
        pol = ws.cell(r, 8).value
        pod = ws.cell(r, 9).value
        penalty_type = ws.cell(r, 15).value
        carrier = ws.cell(r, 4).value
        if not type_ or not carrier:
            continue
        port = pod if type_ == "Destination" else pol
        existing_index[(carrier, type_, port, penalty_type, container_type)] = r

    next_row = ws.max_row + 1
    appended = 0
    updated = 0
    skipped_manual = 0
    today_str = __import__("datetime").date.today().isoformat()

    for doc in docs:
        for mrow in map_doc(doc):
            key = row_key(mrow)
            existing_row = existing_index.get(key)

            if existing_row is not None:
                if ws.cell(existing_row, MANUALLY_REVIEWED_COL).value:
                    skipped_manual += 1
                    continue
                new_end = parse_date_best_effort(mrow.contract_end)
                if new_end is not None:
                    current_end = parse_date_best_effort(ws.cell(existing_row, 13).value)
                    if current_end is None or new_end > current_end:
                        ws.cell(existing_row, 13, mrow.contract_end)
                        note_cell = ws.cell(existing_row, 26)
                        audit = f"[{today_str}] end date auto-refreshed to {mrow.contract_end}"
                        note_cell.value = f"{note_cell.value} | {audit}" if note_cell.value else audit
                        updated += 1
                continue

            r = next_row
            ws.cell(r, 1, mrow.type_)
            ws.cell(r, 2, mrow.container_type)
            ws.cell(r, 3, mrow.hazardous)
            ws.cell(r, 4, mrow.carrier)
            if e_formula:
                ws.cell(r, 5, retarget_formula(e_formula, r))
            ws.cell(r, 6, mrow.shipper)
            ws.cell(r, 7, mrow.por)
            ws.cell(r, 8, mrow.pol)
            ws.cell(r, 9, mrow.pod)
            ws.cell(r, 10, mrow.podel)
            ws.cell(r, 11, mrow.contract_number)
            ws.cell(r, 12, mrow.contract_start)
            ws.cell(r, 13, mrow.contract_end)
            ws.cell(r, 14, mrow.validity_milestone)
            ws.cell(r, 15, mrow.penalty_type)
            ws.cell(r, 16, mrow.currency)
            ws.cell(r, 17, mrow.date_calc)
            if r_formula:
                ws.cell(r, 18, retarget_formula(r_formula, r))
            if s_formula:
                ws.cell(r, 19, retarget_formula(s_formula, r))
            ws.cell(r, 20, mrow.free_time)
            ws.cell(r, 21, mrow.penalty1)
            ws.cell(r, 22, mrow.period1)
            ws.cell(r, 23, mrow.penalty2)
            ws.cell(r, 24, mrow.period2)
            ws.cell(r, 25, mrow.penalty3)
            qa_text = " | ".join(mrow.qa_notes)
            ws.cell(r, 26, qa_text)
            if mrow.qa_notes:
                for col in range(1, 27):
                    ws.cell(r, col).fill = NEEDS_REVIEW_FILL

            existing_index[key] = r  # so a later doc in this same run can match it too
            next_row += 1
            appended += 1

    wb.save(out_path)
    return {
        "rows_appended": appended,
        "rows_end_date_updated": updated,
        "rows_skipped_manually_reviewed": skipped_manual,
        "output": str(out_path),
    }


if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--template", type=Path, required=True,
                     help="Path to BuyCo's Template (6).xlsx")
    ap.add_argument("--index-csv", type=Path, required=True,
                     help="dnd_tariff_index_*.csv from a scraper run "
                          "(re-fetches each PDF's raw text before mapping)")
    ap.add_argument("--out", type=Path, default=Path("dnd_tariffs_mapped.xlsx"))
    args = ap.parse_args()

    import csv as _csv
    from dnd_tariff_scraper import TariffDoc as _TariffDoc, fetch_and_parse_pdf

    docs = []
    with args.index_csv.open(newline="", encoding="utf-8") as fh:
        for row in _csv.DictReader(fh):
            d = _TariffDoc(
                carrier=row["Carrier"], region=row.get("Region", ""),
                country=row["Country"], title=row["Title"], pdf_url=row["PDF URL"],
            )
            fetch_and_parse_pdf(d)
            docs.append(d)

    summary = write_template(args.template, docs, args.out)
    print(summary)
